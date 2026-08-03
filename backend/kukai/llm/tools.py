"""Tool definitions for the LLM — maps to bridge methods."""

from __future__ import annotations

import os
from typing import Any, Optional


def tools_v2_enabled() -> bool:
    """KUKAI_TOOLS_V2=1 turns the merged Tool Palette v2 on (env read at CALL
    time — the KUKAI_EXEC_PIPELINE convention: restart-free operator flips,
    a typo can never activate). Default OFF ⇒ today's palette byte-identical."""
    return os.environ.get("KUKAI_TOOLS_V2", "0") == "1"


def tools_v3_enabled() -> bool:
    """KUKAI_TOOLS_V3=1 (on top of v2) folds the read cluster get_model_info +
    inspect into ONE inspect(scope=model|element) tool. Default OFF ⇒ v2 palette
    unchanged, so it can ship dark and be smoke-tested before the fleet sees it."""
    return os.environ.get("KUKAI_TOOLS_V3", "0") == "1"


def expects_contract_enabled() -> bool:
    """KUKAI_EXPECTS_CONTRACT=1 turns the expects postcondition contract on
    (IQ moment N2): the execute_revit_code schema gains the optional `expects`
    field and the exec pipeline witnesses it with read-only before/after
    category-count probes. Env read at CALL time (project flag convention).
    Default OFF ⇒ schema byte-identical, zero probe roundtrips."""
    return os.environ.get("KUKAI_EXPECTS_CONTRACT", "0") == "1"


def _expects_schema_fragment() -> dict[str, Any]:
    """The optional `expects` property for execute_revit_code — empty dict when
    the flag is OFF (splatted into the properties literal ⇒ not a token more).

    The op enum / category / count contract mirrors
    kukai.will.evaluator.parse_expects (lenient: a wrong category or count can
    only ever degrade the verdict to `unverifiable`, never error the call)."""
    if not expects_contract_enabled():
        return {}
    return {
        "expects": {
            "type": "object",
            "description": (
                "ОПЦИОНАЛЬНО для операций ЗАПИСИ: контракт результата — что код "
                "должен изменить в модели. Система сама проверит подсчётом "
                "элементов категории до/после выполнения (модель не трогает). "
                "Пример: {\"op\": \"create\", \"category\": \"OST_Walls\", \"count\": 5}. "
                "Для чтения не передавай."
            ),
            "properties": {
                "op": {
                    "type": "string",
                    "enum": ["create", "modify", "delete"],
                    "description": "Тип изменения: create (создание), modify (изменение), delete (удаление).",
                },
                "category": {
                    "type": "string",
                    "description": (
                        "Имя BuiltInCategory затрагиваемых элементов, напр. "
                        "'OST_Walls', 'OST_Doors', 'OST_Windows'."
                    ),
                },
                "count": {
                    "type": "integer",
                    "minimum": 0,
                    "description": "Сколько элементов будет создано/удалено.",
                },
            },
        },
    }


# Tools hidden when the active doc is a family (.rfa) — Excel/VOR/Schedule
# pipelines are project-only and add noise + token cost in family editor mode.
_NON_FAMILY_TOOLS = frozenset({
    "generate_report",
    "modify_excel",
    "excel_script",
    "price_vor",
    "lookup_norm",
    "import_cad",
    "export_sheets_pdf",
    "apply_revit_write",
    "query_model",
})


def get_tool_definitions(
    module_registry=None,
    context: Optional[Any] = None,
) -> list[dict[str, Any]]:
    """Return OpenAI-format tool definitions for the LLM.

    These tools map to bridge methods:
    - execute_revit_code -> bridge.execute
    - get_model_info -> bridge.context
    - select_elements -> bridge.select
    - highlight_elements -> bridge.highlight

    If module_registry is provided, appends tool definitions from
    loaded KUKAI modules (audit, commands, rascenka, etc.).
    """
    hardcoded = [
        {
            "type": "function",
            "function": {
                "name": "execute_revit_code",
                "description": (
                    "Execute work with Revit model. "
                    "PHASE 1 (USE_REVIT_CODER=1): pass `task` (English description, no pseudocode). "
                    "Backend invokes specialized code generator, validates via Roslyn, executes via Bridge. "
                    "PHASE 0 (USE_REVIT_CODER=0, current default): pass `code` (literal C#). "
                    "Code runs in: public static object Execute(Document doc, UIDocument uidoc). "
                    "Available namespaces: Autodesk.Revit.*, System, System.Linq, "
                    "System.Collections.Generic, System.Text. "
                    "For write operations, code must wrap in a Transaction."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "task": {
                            "type": "string",
                            "description": (
                                "Plain-English description of what the code should do. "
                                "Describe WHAT, not HOW. No pseudocode, no API names. "
                                "Example: 'Count walls in active view, grouped by wall type.' "
                                "Required when USE_REVIT_CODER=1."
                            ),
                        },
                        "code": {
                            "type": "string",
                            "description": (
                                "Literal C# code to execute. "
                                "Required when USE_REVIT_CODER=0 (current default). "
                                "Must return a value. Do NOT use System.IO, System.Net, "
                                "System.Diagnostics.Process, or any file/network operations."
                            ),
                        },
                        "allow_destructive": {
                            "type": "boolean",
                            "description": (
                                "Требуется при doc.Delete."
                            ),
                        },
                        "model_context": {
                            "type": "object",
                            "description": (
                                "Optional Revit model state. Backend auto-fills relevant pieces. "
                                "Pass only fields that affect WHICH API to use or WHICH elements to operate on."
                            ),
                            "properties": {
                                "revit_version": {"type": "string", "description": "'2024' | '2025' | '2026'"},
                                "active_view_id": {"type": "integer", "description": "Active view ElementId"},
                                "active_view_type": {"type": "string", "description": "FloorPlan|Section|ThreeD|..."},
                                "project_units": {"type": "string", "description": "Millimeters|Feet|..."},
                                "selected_element_ids": {
                                    "type": "array",
                                    "items": {"type": "integer"},
                                    "description": "Pass when user said 'these' / 'selected'.",
                                },
                            },
                        },
                        "previous_code": {
                            "type": "string",
                            "description": (
                                "When user asks to modify the LAST code generated in this conversation, "
                                "pass that previous code here. Otherwise omit."
                            ),
                        },
                        "estimated_elements": {
                            "type": "integer",
                            "description": (
                                "Estimated number of elements this operation will affect. "
                                "Used to set an appropriate timeout. "
                                "For read operations, omit this. For write operations, "
                                "estimate how many elements will be modified."
                            ),
                        },
                        # expects postcondition contract (KUKAI_EXPECTS_CONTRACT,
                        # IQ N2) — {} when OFF ⇒ schema byte-identical.
                        **_expects_schema_fragment(),
                    },
                    # NOTE: "required" intentionally omitted — `task` OR `code` is required,
                    # validated at runtime in client.py:_execute_tool. JSON Schema can't
                    # express XOR cleanly.
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "get_model_info",
                "description": (
                    "Get information about the currently open Revit model: "
                    "document name, categories with element counts, levels, "
                    "current view, selected elements, phase, and warnings count."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {},
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "get_model_details",
                "description": (
                    "Get the FULL detailed passport of the open Revit model when the brief "
                    "context is not enough: family types per category, shared/project parameter "
                    "tables with sample values, grids/phases/rooms, views/sheets/schedules and "
                    "project standards. Call this BEFORE writing code that references specific "
                    "parameters, family types or schedules. Optionally request a single section "
                    "to keep the answer small."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "section": {
                            "type": "string",
                            "enum": ["full", "structure", "parameters", "spatial", "views", "standards"],
                            "description": (
                                "Which part to fetch: 'full' (everything), 'structure' "
                                "(family types), 'parameters' (shared/project params + sample "
                                "values), 'spatial' (grids/phases/rooms/bbox), 'views' "
                                "(views/sheets/schedules), 'standards' (units/naming/"
                                "classification). Default 'full'."
                            ),
                        },
                    },
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "select_elements",
                "description": (
                    "Select elements in the Revit viewport by their Element IDs. "
                    "The elements will be highlighted in the model view."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "element_ids": {
                            "type": "array",
                            "items": {"type": "integer"},
                            "description": "List of Revit Element IDs to select.",
                        },
                    },
                    "required": ["element_ids"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "highlight_elements",
                "description": (
                    "Highlight elements with a color override in the current Revit view. "
                    "Useful for visually marking problematic or relevant elements."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "element_ids": {
                            "type": "array",
                            "items": {"type": "integer"},
                            "description": "List of Revit Element IDs to highlight.",
                        },
                        "color": {
                            "type": "object",
                            "properties": {
                                "r": {"type": "integer", "description": "Red (0-255)"},
                                "g": {"type": "integer", "description": "Green (0-255)"},
                                "b": {"type": "integer", "description": "Blue (0-255)"},
                            },
                            "description": "RGB color for highlighting. Default: red.",
                        },
                        "clear_previous": {
                            "type": "boolean",
                            "description": "Clear previous highlights before applying new ones. Default: true.",
                        },
                    },
                    "required": ["element_ids"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "generate_report",
                "description": (
                    "Generate an Excel (.xlsx) report file from structured data. "
                    "RULES (read every time before calling): "
                    "1) Sheet 1 = raw data only. ALL rows, no formulas, no filtering. This is the untouchable source. "
                    "2) Pass ALL data from execute_revit_code — do NOT truncate, limit, or .Take(). System handles 10K+ rows. "
                    "3) Column keys MUST be human-readable: Name, Level, Area_m2, Volume_m3. NEVER use col_0, col_1, or generic names. "
                    "4) If user wants summaries/pivots — use 'operations' param or call excel_script AFTER to add new sheets. "
                    "5) NEVER call this without first collecting data via execute_revit_code. "
                    "6) For multi-category: use 'sheets' param with one sheet per category."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "data": {
                            "type": "array",
                            "items": {"type": "object"},
                            "description": (
                                "List of row objects. Each object's keys become column headers. "
                                "Example: [{\"Name\": \"Wall-1\", \"Length_m\": 5.2}, ...]"
                            ),
                        },
                        "filename": {
                            "type": "string",
                            "description": "Output filename (without path). Default: report.xlsx",
                        },
                        "sheet_name": {
                            "type": "string",
                            "description": "Excel sheet name. Default: Report",
                        },
                        "sort_by": {
                            "type": "string",
                            "description": (
                                "Column name to sort data by before generating the file. "
                                "Example: 'Area', 'Level', 'Name'."
                            ),
                        },
                        "sort_order": {
                            "type": "string",
                            "enum": ["asc", "desc"],
                            "description": "Sort order: 'asc' (ascending, default) or 'desc' (descending).",
                        },
                        "operations": {
                            "type": "array",
                            "items": {"type": "object"},
                            "description": (
                                "Advanced Excel operations to apply after creating the file. "
                                "Each operation is an object with 'type' and parameters. "
                                "Supported types: "
                                "'sort' (column, order), "
                                "'add_totals' (columns: list of column names), "
                                "'add_formula_column' (name, formula — use {row} for row number), "
                                "'conditional_format' (column, operator, value, color), "
                                "'pivot' (rows, values, aggfunc: sum/count/avg, sheet_name). "
                                "Example: [{\"type\": \"pivot\", \"rows\": \"Level\", \"values\": \"Area\", \"aggfunc\": \"sum\"}]"
                            ),
                        },
                        "sheets": {
                            "type": "array",
                            "items": {
                                "type": "object",
                                "properties": {
                                    "name": {"type": "string", "description": "Sheet name"},
                                    "data": {"type": "array", "items": {"type": "object"}, "description": "Rows for this sheet"},
                                },
                            },
                            "description": "Multiple sheets. If provided, 'data' and 'sheet_name' are ignored. Each item: {name, data}.",
                        },
                    },
                    "required": ["data"],
                },
            },
        },
        # --- File import/export tools ---
        {
            "type": "function",
            "function": {
                "name": "export_view",
                "description": (
                    "ПОСМОТРЕТЬ на текущий вид Revit: делает снимок и ВОЗВРАЩАЕТ ТЕБЕ "
                    "изображение — ты увидишь его своими глазами следующим сообщением. "
                    "Вызывай, когда нужно убедиться в результате: закончил этаж, собрал "
                    "узел, доделал задачу — посмотри, получилось ли задуманное и нет ли "
                    "ошибок размещения. Не нужно смотреть после каждого мелкого шага. "
                    "Также используй, если пользователь просит сохранить/экспортировать вид."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "filename": {
                            "type": "string",
                            "description": "Output filename (no path). Example: 'plan_floor1.png'",
                        },
                        "format": {
                            "type": "string",
                            "enum": ["png", "jpg"],
                            "description": "Image format. Default: PNG.",
                        },
                    },
                    "required": [],
                },
            },
        },
        # import_cad REMOVED from the palette 2026-07-17 (operator call): a thin
        # wrapper over doc.Import (~5 trivial lines) that returns only an
        # element_id, not the useful curves — 0 successful calls in 30d, worst of
        # both worlds (occupies a slot, gives ~no abstraction). Rare disk-insert →
        # execute_revit_code; a DWG already an underlay needs no import at all.
        # Dispatch handler kept as a no-op alias for cached conversations.
        # --- PDF Export & File Delivery tools ---
        {
            "type": "function",
            "function": {
                "name": "export_sheets_pdf",
                "description": (
                    "Export Revit ViewSheets to PDF files via doc.Export(). "
                    "Returns file paths. Call send_local_file for each PDF to deliver to user."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "sheet_ids": {
                            "type": "array",
                            "items": {"type": "integer"},
                            "description": "ElementId list of sheets to export. Empty = export all sheets.",
                        },
                        "combine": {
                            "type": "boolean",
                            "description": "Combine all sheets into one PDF. Default: false.",
                        },
                        "quality": {
                            "type": "string",
                            "enum": ["draft", "standard", "high"],
                            "description": "PDF quality: draft (150dpi), standard (300dpi), high (600dpi). Default: standard.",
                        },
                    },
                    "required": [],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "send_local_file",
                "description": (
                    "Send a local file to user for download. "
                    "Use AFTER Revit exported files (e.g. PDF via export_sheets_pdf). "
                    "File must exist on server."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "file_path": {
                            "type": "string",
                            "description": "Absolute path to file on local machine.",
                        },
                        "filename": {
                            "type": "string",
                            "description": "Display filename for download. Example: 'AR-01_Floor_Plan.pdf'",
                        },
                    },
                    "required": ["file_path"],
                },
            },
        },
        # --- Extended tools ---
        {
            "type": "function",
            "function": {
                "name": "apply_revit_write",
                "description": (
                    "Safe Revit model write operations. Use INSTEAD of execute_revit_code "
                    "for simple ops: set parameter, create schedule, hide/isolate, rename, "
                    "delete, copy, move. Faster and safer than raw code."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "operation": {
                            "type": "string",
                            "enum": [
                                "set_parameter", "create_schedule", "hide_or_isolate",
                                "rename_entities", "delete_elements", "copy_elements",
                                "move_elements",
                            ],
                            "description": "Write operation type.",
                        },
                        "category": {
                            "type": "string",
                            "description": "Категория элементов (walls, floors, columns, etc.)",
                        },
                        "element_ids": {
                            "type": "array",
                            "items": {"type": "integer"},
                            "description": "ElementId list (empty = use current working set)",
                        },
                        "parameter_name": {
                            "type": "string",
                            "description": "Parameter name for set_parameter.",
                        },
                        "value": {
                            "type": "string",
                            "description": "Value to set.",
                        },
                        "schedule_name": {
                            "type": "string",
                            "description": "Schedule name for create_schedule.",
                        },
                        "schedule_fields": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Field names to add to schedule.",
                        },
                        "view_action": {
                            "type": "string",
                            "enum": ["hide", "isolate"],
                            "description": "Action for hide_or_isolate.",
                        },
                        "new_name": {
                            "type": "string",
                            "description": "New name for rename_entities.",
                        },
                        "rename_mode": {
                            "type": "string",
                            "enum": ["exact", "prefix", "suffix"],
                            "description": "Rename mode.",
                        },
                        "offset_x": {
                            "type": "number",
                            "description": "X offset in millimeters (for copy/move).",
                        },
                        "offset_y": {
                            "type": "number",
                            "description": "Y offset in millimeters (for copy/move).",
                        },
                        "offset_z": {
                            "type": "number",
                            "description": "Z offset in millimeters (for copy/move).",
                        },
                    },
                    "required": ["operation"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "query_model",
                "description": (
                    "Найти / посчитать / агрегировать / выделить элементы по "
                    "ДЕКЛАРАТИВНОМУ фильтру — БЕЗ написания C#. Надёжнее "
                    "execute_revit_code для поиска: бэкенд исполняет проверенный "
                    "version-safe шаблон. Фильтр по категории/типу/параметру → "
                    "count | ids | площадь/объём | группировка, и опц. действие "
                    "select/isolate/highlight. ВАЖНО: точные имена типов бери из "
                    "паспорта модели (раздел СЛОВАРЬ: монолит/перегородки/наружные "
                    "→ точные имена), передавай в type_contains или type_names."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "category": {
                            "type": "string",
                            "description": "Категория элементов (стены/двери/walls/floors/...) — RU/EN алиас.",
                        },
                        "type_contains": {
                            "type": "string",
                            "description": "Подстрока имени ТИПА (регистронезависимо), напр. 'Монолит'.",
                        },
                        "type_names": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Точные имена типов (из СЛОВАРЯ паспорта) — самый надёжный фильтр.",
                        },
                        "param": {
                            "type": "object",
                            "description": "Фильтр по параметру экземпляра.",
                            "properties": {
                                "name": {"type": "string", "description": "Имя параметра."},
                                "op": {
                                    "type": "string",
                                    "enum": ["empty", "not_empty", "eq", "contains", "gt", "lt"],
                                },
                                "value": {"type": "string", "description": "Значение (для eq/contains/gt/lt)."},
                            },
                        },
                        "function": {
                            "type": "string",
                            "enum": ["exterior", "interior", "foundation", "retaining", "coreshaft", "soffit"],
                            "description": "Функция стены по РЕАЛЬНОМУ свойству WallType.Function "
                                           "(наружная=exterior, внутренняя/перегородка=interior). Точнее имени типа.",
                        },
                        "width_mm": {
                            "type": "object",
                            "description": "Фильтр по РЕАЛЬНОЙ толщине стены (WallType.Width), в мм.",
                            "properties": {
                                "op": {"type": "string", "enum": ["gt", "lt", "eq", "range"]},
                                "value": {"type": "number"},
                                "value2": {"type": "number", "description": "Верхняя граница для op=range."},
                            },
                        },
                        "layer_material_contains": {
                            "type": "string",
                            "description": "Подстрока имени МАТЕРИАЛА слоя конструкции (стены/перекрытия). "
                                           "Монолит/ЖБ → 'Железобетон' (исключает блок/ЯБ). Точнее имени типа.",
                        },
                        "level": {
                            "type": "string",
                            "description": "Подстрока имени уровня элемента, напр. '03' для «3 этаж» (= 03_К1.2).",
                        },
                        "selected": {
                            "type": "boolean",
                            "description": "Работать по текущему выделению, а не по всей модели.",
                        },
                        "return": {
                            "type": "string",
                            "enum": ["count", "ids", "aggregate", "group", "coverage"],
                            "description": "Что вернуть (по умолчанию count). coverage = "
                                           "заполненность param.name по категории за ОДИН вызов "
                                           "(total/filled/empty/empty_ids) — для нормоконтроля.",
                        },
                        "aggregate": {
                            "type": "array",
                            "items": {"type": "string", "enum": ["count", "area_m2", "volume_m3"]},
                            "description": "Метрики для return=aggregate.",
                        },
                        "group_by": {
                            "type": "string",
                            "enum": ["type", "level"],
                            "description": ("Группировка для return=group: по типу или "
                                            "по этажу (aggregate area_m2/volume_m3 "
                                            "считается на каждый уровень)."),
                        },
                        "action": {
                            "type": "string",
                            "enum": ["none", "select", "isolate", "highlight"],
                            "description": "Действие над найденными (по умолчанию none).",
                        },
                        "limit": {
                            "type": "integer",
                            "description": "Предел числа id для return=ids (по умолчанию 1000).",
                        },
                    },
                    "required": [],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "process_uploaded_file",
                "description": (
                    "Process uploaded file (Excel, PDF, Word, JSON, image). "
                    "Extract content for analysis."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "file_id": {
                            "type": "string",
                            "description": "File ID from upload.",
                        },
                        "extraction_mode": {
                            "type": "string",
                            "enum": ["text", "structured", "ocr"],
                            "description": "Extraction mode. Default: text.",
                        },
                    },
                    "required": ["file_id"],
                },
            },
        },
        # add_user_note REMOVED from the palette 2026-07-17 (operator call): low
        # value (8 calls/30d, unreliable trigger), one fewer tool for the model
        # to weigh. The dispatch handler stays as a harmless no-op alias so any
        # cached/old conversation that still emits it does not error.
        {
            "type": "function",
            "function": {
                "name": "lookup_norm",
                "description": "Search Russian building codes and normative documents (СП, ГОСТ, ПУЭ, ФЗ). "
                "Returns relevant paragraphs and clauses from official regulatory documents. "
                "Use this tool when the user asks about building code requirements, norms, regulations, "
                "or needs specific clause references. This is a FAST lookup — just call it with a query "
                "and get results immediately. No reasoning needed.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "query": {
                            "type": "string",
                            "description": "Search query in Russian (e.g., 'минимальная ширина коридора', "
                            "'огнестойкость перекрытий', 'допустимый ток кабеля')",
                        },
                        "limit": {
                            "type": "integer",
                            "description": "Max results (default 5, max 15)",
                        },
                    },
                    "required": ["query"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "price_vor",
                "description": (
                    "Auto-price a ВОР (Bill of Quantities) Excel file. "
                    "Parses all positions, matches ГЭСН/ФЕР codes, calculates cost "
                    "with resource breakdown via multiagent expert pipeline, generates "
                    "priced Excel with confidence colors. "
                    "Use when user asks to price ВОР, find расценки, or sends /расценка."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "file_content": {
                            "type": "string",
                            "description": "Uploaded ВОР Excel file content. Use content from context if uploaded via /chat/file.",
                        },
                    },
                    "required": [],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "modify_excel",
                "description": (
                    "Apply operations to the last generated Excel report. "
                    "Use AFTER generate_report. Adds NEW sheets — never modifies raw data on Sheet 1. "
                    "Pivot creates a new sheet with SUMIF/COUNTIF formulas referencing the data sheet. "
                    "Example: user says 'add a pivot by Level to that report'."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "operations": {
                            "type": "array",
                            "items": {"type": "object"},
                            "description": (
                                "Operations to apply. Each has 'type' + params. Types: "
                                "'sort' (column, order), "
                                "'auto_filter', "
                                "'freeze_header', "
                                "'column_widths', "
                                "'header_format' (color — hex without #), "
                                "'add_totals' (columns: list), "
                                "'add_formula_column' (name, formula — {row} = row number), "
                                "'conditional_format' (column, operator: greaterThan/lessThan/equal, value, color), "
                                "'pivot' (rows, values, aggfunc: sum/count/avg, sheet_name)."
                            ),
                        },
                        "filename": {
                            "type": "string",
                            "description": "Output filename. Default: same as original with '_modified' suffix.",
                        },
                    },
                    "required": ["operations"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "excel_script",
                "description": (
                    "Execute openpyxl Python script on the current Excel file. "
                    "RULES: "
                    "1) NEVER modify Sheet 1 (raw data). ALWAYS create NEW sheets for analysis. "
                    "2) Use FORMULAS referencing Sheet 1: =SUMIF('Данные'!A:A, ...), =COUNTIF(), =IF(). "
                    "   NEVER calculate values in Python — write Excel formulas so the file is 'smart'. "
                    "3) The file has ALL rows (thousands). Use ws.max_row to iterate — never hardcode row counts. "
                    "4) Available: openpyxl (all submodules), math, datetime, decimal. "
                    "5) Variable `wb` is pre-loaded Workbook. Create new sheets: wb.create_sheet('Сводная'). "
                    "Use for: summaries, pivots, percentage columns, conditional formatting, charts."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "script": {
                            "type": "string",
                            "description": (
                                "Python script using openpyxl. "
                                "Variable `wb` is the pre-loaded Workbook. "
                                "Modify wb in-place — changes are auto-saved. "
                                "Example:\n"
                                "from openpyxl.styles import Font, PatternFill\n"
                                "from openpyxl.utils import get_column_letter\n"
                                "ws = wb.active\n"
                                "# Add percentage column\n"
                                "ws.cell(1, 4, value='% от общей')\n"
                                "for r in range(2, ws.max_row + 1):\n"
                                "    ws.cell(r, 4, value=f'=B{r}/SUM(B$2:B${ws.max_row})*100')\n"
                            ),
                        },
                        "filename": {
                            "type": "string",
                            "description": "Output filename. Default: original name with '_scripted' suffix.",
                        },
                    },
                    "required": ["script"],
                },
            },
        },
    ]

    # ─── Phase 1 (revit-coder pilot) tool filtering ───
    # When USE_REVIT_CODER=1, exclude heavy pipelines and duplicates.
    # See docs/superpowers/specs/2026-05-01-revit-coder-integration-design.md
    from kukai.config import USE_REVIT_CODER, DISABLED_TOOLS_REVIT_CODER_MODE

    # Filter #1: hardcoded tools
    if USE_REVIT_CODER:
        hardcoded = [
            t for t in hardcoded
            if t["function"]["name"] not in DISABLED_TOOLS_REVIT_CODER_MODE
        ]

    # Filter #2: module_registry tools (audit, vor, scheduling)
    if module_registry is not None:
        module_tools = module_registry.get_all_tools()
        hardcoded_names = {t["function"]["name"] for t in hardcoded}
        for tool_def in module_tools:
            if tool_def.name in hardcoded_names:
                continue
            if USE_REVIT_CODER and tool_def.name in DISABLED_TOOLS_REVIT_CODER_MODE:
                continue  # CRITICAL: same filter for module tools
            hardcoded.append(tool_def.to_openai_dict())

    # ─── Family-editor mode tool gating (V2) ───────────────────────────────
    # When the active doc is a FAMILY (.rfa):
    #   - HIDE project-only tools (Excel, VOR, schedules, lookups) — irrelevant noise
    #   - ADD 10 purpose-built family_* tools so Gemini composes deterministic
    #     server-side templates instead of writing brittle parametric C# from scratch
    is_family_editor = bool(context and getattr(context, "is_family_editor", False))
    if is_family_editor:
        hardcoded = [
            t for t in hardcoded
            if t["function"]["name"] not in _NON_FAMILY_TOOLS
        ]
        hardcoded.extend(_get_family_tool_definitions())

    # ─── query_model (G3) gated rollout ───────────────────────────────────
    # Additive declarative discovery tool; expose only when KUKAI_QUERY_MODEL=1.
    try:
        from kukai.config import get_settings as _gs
        _qm_on = _gs().query_model
    except Exception:
        _qm_on = False
    if not _qm_on:
        hardcoded = [t for t in hardcoded if t["function"]["name"] != "query_model"]

    # ─── Wave 1: `inspect` drill verb — gated by KUKAI_PERCEPTION ──────────────
    try:
        from kukai.config import PERCEPTION as _perc
    except Exception:
        _perc = False
    if _perc and not any(t["function"]["name"] == "inspect" for t in hardcoded):
        hardcoded.append({
            "type": "function",
            "function": {
                "name": "inspect",
                "description": (
                    "Раскрыть ВСЕ свойства ОДНОГО элемента по id (как клик по элементу "
                    "в Ревите → палитра): категория, тип, уровень, все непустые параметры. "
                    "Надёжнее/быстрее execute_revit_code для осмотра элемента. id брать из "
                    "query_model return=ids или из выделения."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "element_id": {
                            "type": "string",
                            "description": "Числовой id элемента (напр. из query_model return=ids).",
                        },
                    },
                    "required": ["element_id"],
                },
            },
        })

    # ─── ARCHIVED 2026-06-10 (operator: archive VOR): drop the VOR pricing tool ──
    hardcoded = [t for t in hardcoded if t["function"]["name"] != "price_vor"]

    # ─── Temporarily FROZEN tools (global, reversible via KUKAI_FROZEN_TOOLS) ──
    try:
        from kukai.config import FROZEN_TOOLS as _frozen
    except Exception:
        _frozen = set()
    if _frozen:
        hardcoded = [t for t in hardcoded if t["function"]["name"] not in _frozen]

    # ─── Tool Palette v2 (merged tools) — gated by KUKAI_TOOLS_V2 ─────────────
    # Flag OFF (default) → this is a no-op and the list above is byte-identical
    # to the pre-v2 build. Flag ON → overlapping tools are merged (17 → 12):
    # select+highlight→show_elements, modify_excel+excel_script→edit_excel,
    # export_view+export_sheets_pdf(+send_local_file folded)→export, and
    # get_model_details is absorbed into query_model(scope="details") whose
    # 14 loose props collapse into categories[]+typed filter+scope. Placed
    # AFTER every mode/flag filter (family/coder/frozen stay authoritative on
    # the OLD names) and BEFORE create_element injection (which targets
    # apply_revit_write — present unchanged in both palettes).
    if tools_v2_enabled():
        hardcoded = _apply_v2_palette(hardcoded)

    # ─── create_element (declarative creation) — gated by KUKAI_CREATE_ELEMENT ──
    # Flag OFF (default) → the op is ABSENT from the schema entirely (not merely
    # rejected at call time), so flag-off turns are byte-identical to the
    # pre-create_element build (same additive-gating contract as query_model
    # above). Single source of truth for the enum member + the nested `element`
    # schema lives in kukai.write.create_element (design 2026-07-04). Placed
    # after ALL filters so family-editor mode (apply_revit_write hidden) and
    # frozen-tools stay authoritative — injection no-ops when the tool is absent.
    try:
        from kukai.write.create_element import (
            create_element_enabled as _ce_enabled,
            inject_create_element_schema as _ce_inject,
        )
        if _ce_enabled():
            _ce_inject(hardcoded)
    except Exception:  # noqa: BLE001 — schema injection must never break the tool list
        pass

    # ─── revit_ir (KIR stage 2) — gated by KUKAI_KIR_TOOL=stage2 + admin device ──
    # Same additive-gating contract as create_element: gate-off turns are
    # byte-identical; the gate itself lives in kukai.ir.serving (fail-closed).
    try:
        from kukai.ir.serving import (
            revit_ir_enabled as _kir_enabled,
            inject_revit_ir_schema as _kir_inject,
        )
        if _kir_enabled():
            _kir_inject(hardcoded)
    except Exception:  # noqa: BLE001 — schema injection must never break the tool list
        pass

    return hardcoded


# ─────────────────────────────────────────────────────────────────────────────
# Tool Palette v2 (KUKAI_TOOLS_V2) — merged tools, thinned params
#
# Design (task 2026-07-04): DeepSeek grounds tool NAMES fine but fails fat
# PARAMS (query_model's 14 loose props = worst case, proven by the operator's
# Claude-vs-DeepSeek study), and overlapping read tools cause churn (a simple
# "count walls+doors+windows" burned query_model×3 + execute = 6 bridge
# round-trips). v2 merges overlap and nests semantic filters into ONE compact
# typed object. Dispatch compat: kukai.llm.tool_handlers.palette_v2 LOWERS the
# new names onto the existing (unchanged) legacy handlers, and the old names
# remain valid aliases at the dispatch layer — an LLM continuing an old
# conversation or a cached prompt never breaks.
# ─────────────────────────────────────────────────────────────────────────────

# Old-name → merged-tool map (single source of truth for the transform below
# and for the deprecation log in palette_v2).
V2_MERGES: dict[str, str] = {
    "select_elements": "show_elements",
    "highlight_elements": "show_elements",
    "modify_excel": "edit_excel",
    "excel_script": "edit_excel",
    "export_view": "export",
    "export_sheets_pdf": "export",
    "send_local_file": "export",       # delivery folded into export(deliver=…)
    "get_model_details": "query_model",  # absorbed as scope="details"
}

V2_NEW_TOOLS = frozenset({"show_elements", "edit_excel", "export"})


def _apply_v2_palette(tools: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Order-preserving merge: each merged tool takes the slot of its FIRST
    surviving constituent. Merges only fire when a constituent actually
    survived the upstream mode/flag filters, so no capability ever silently
    vanishes (e.g. family-editor mode hides query_model → get_model_details
    stays; send_local_file is only folded when an export tool exists)."""
    names = {t["function"]["name"] for t in tools}
    has_export_constituent = bool({"export_view", "export_sheets_pdf"} & names)
    absorb_details = "query_model" in names and "get_model_details" in names
    # v3: fold get_model_info + inspect → inspect(scope). Only when BOTH survived
    # the upstream filters, so no read capability silently vanishes.
    absorb_gmi = tools_v3_enabled() and "get_model_info" in names and "inspect" in names
    out: list[dict[str, Any]] = []
    emitted: set[str] = set()
    for t in tools:
        n = t["function"]["name"]
        if absorb_gmi and n in ("get_model_info", "inspect"):
            if "inspect" not in emitted:
                out.append(_inspect_v3_def())
                emitted.add("inspect")
        elif n in ("select_elements", "highlight_elements"):
            if "show_elements" not in emitted:
                out.append(_show_elements_def())
                emitted.add("show_elements")
        elif n in ("modify_excel", "excel_script"):
            if "edit_excel" not in emitted:
                out.append(_edit_excel_def())
                emitted.add("edit_excel")
        elif n in ("export_view", "export_sheets_pdf"):
            if "export" not in emitted:
                out.append(_export_def())
                emitted.add("export")
        elif n == "send_local_file":
            if not has_export_constituent:
                out.append(t)  # nothing absorbs it here — keep the capability
        elif n == "get_model_details":
            if not absorb_details:
                out.append(t)
        elif n == "query_model":
            out.append(_query_model_v2_def())
        else:
            out.append(t)
    return out


def _inspect_v3_def() -> dict[str, Any]:
    """get_model_info + inspect → ONE tool (scope enum). scope='model' — обзор
    всей модели; scope='element' + element_id — все свойства одного элемента."""
    return {
        "type": "function",
        "function": {
            "name": "inspect",
            "description": (
                "Осмотреть модель или элемент. scope='model' — сводка по открытой "
                "модели (уровни, категории, счётчики; вызывай БЕЗ element_id). "
                "scope='element' — раскрыть ВСЕ свойства ОДНОГО элемента по id "
                "(категория, тип, уровень, все непустые параметры; как клик по "
                "элементу в Ревите). Надёжнее/быстрее execute_revit_code для осмотра."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "scope": {
                        "type": "string",
                        "enum": ["model", "element"],
                        "description": "'model' — обзор модели; 'element' — один элемент по id.",
                    },
                    "element_id": {
                        "type": "string",
                        "description": "Числовой id элемента (для scope='element'; из query_model return=ids).",
                    },
                },
                "required": ["scope"],
            },
        },
    }


def _show_elements_def() -> dict[str, Any]:
    """select_elements + highlight_elements → ONE tool (mode enum)."""
    return {
        "type": "function",
        "function": {
            "name": "show_elements",
            "description": (
                "Показать элементы в Revit по их ElementId: mode='select' — "
                "выделить в модели (default); mode='highlight' — подсветить "
                "цветом в текущем виде (пометить проблемные/найденные)."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "element_ids": {
                        "type": "array",
                        "items": {"type": "integer"},
                        "description": "Revit ElementId list.",
                    },
                    "mode": {
                        "type": "string",
                        "enum": ["select", "highlight"],
                        "description": "select (default) | highlight.",
                    },
                    "color": {
                        "type": "object",
                        "properties": {
                            "r": {"type": "integer"},
                            "g": {"type": "integer"},
                            "b": {"type": "integer"},
                        },
                        "description": "RGB 0-255 для highlight. Default: red.",
                    },
                    "clear_previous": {
                        "type": "boolean",
                        "description": "Сбросить прежнюю подсветку (default true).",
                    },
                },
                "required": ["element_ids"],
            },
        },
    }


def _edit_excel_def() -> dict[str, Any]:
    """modify_excel + excel_script → ONE tool (script = general case,
    operations = simple-ops fast path)."""
    return {
        "type": "function",
        "function": {
            "name": "edit_excel",
            "description": (
                "Доработать ПОСЛЕДНИЙ сгенерированный Excel-файл (use AFTER "
                "generate_report). Общий случай — `script`: Python/openpyxl, "
                "переменная `wb` = открытый Workbook, изменения автосохраняются. "
                "ПРАВИЛА: НИКОГДА не менять Лист 1 (сырые данные) — только "
                "НОВЫЕ листы (wb.create_sheet); пиши ФОРМУЛЫ со ссылками на "
                "данные (=SUMIF/COUNTIF/IF), не значения; ws.max_row вместо "
                "хардкода числа строк. Быстрый путь для типовых операций без "
                "кода — `operations`: sort, auto_filter, freeze_header, "
                "column_widths, header_format, add_totals, add_formula_column, "
                "conditional_format, pivot (rows, values, aggfunc)."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "script": {
                        "type": "string",
                        "description": (
                            "Python/openpyxl скрипт (общий случай). `wb` — "
                            "pre-loaded Workbook. Доступны openpyxl, math, "
                            "datetime, decimal."
                        ),
                    },
                    "operations": {
                        "type": "array",
                        "items": {"type": "object"},
                        "description": (
                            "Типовые операции без кода (fast path). Каждая: "
                            "{'type': …, параметры}. Пример: [{\"type\": \"pivot\", "
                            "\"rows\": \"Level\", \"values\": \"Area\", \"aggfunc\": \"sum\"}]"
                        ),
                    },
                    "filename": {
                        "type": "string",
                        "description": "Имя выходного файла. Default: суффикс к исходному.",
                    },
                },
                "required": [],
            },
        },
    }


def _export_def() -> dict[str, Any]:
    """export_view + export_sheets_pdf → ONE tool; send_local_file delivery
    folded into deliver=true (default) — no follow-up call needed."""
    return {
        "type": "function",
        "function": {
            "name": "export",
            "description": (
                "Экспорт из Revit + автоматическая доставка файлов пользователю "
                "(deliver=true по умолчанию — send_local_file вызывать НЕ нужно). "
                "what='view' — текущий вид в PNG/JPG; what='sheets_pdf' — листы "
                "(ViewSheets) в PDF через doc.Export()."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "what": {
                        "type": "string",
                        "enum": ["view", "sheets_pdf"],
                        "description": "Что экспортировать.",
                    },
                    "filename": {
                        "type": "string",
                        "description": "Имя файла для what='view'. Пример: 'plan_floor1.png'.",
                    },
                    "format": {
                        "type": "string",
                        "enum": ["png", "jpg"],
                        "description": "Формат изображения для what='view'. Default: png.",
                    },
                    "sheet_ids": {
                        "type": "array",
                        "items": {"type": "integer"},
                        "description": "ElementId листов для what='sheets_pdf'. Пусто = все листы.",
                    },
                    "combine": {
                        "type": "boolean",
                        "description": "sheets_pdf: объединить в один PDF. Default: false.",
                    },
                    "quality": {
                        "type": "string",
                        "enum": ["draft", "standard", "high"],
                        "description": "sheets_pdf: 150/300/600 dpi. Default: standard.",
                    },
                    "deliver": {
                        "type": "boolean",
                        "description": "Отправить файл(ы) пользователю (default true).",
                    },
                },
                "required": ["what"],
            },
        },
    }


def _query_model_v2_def() -> dict[str, Any]:
    """query_model v2: multi-category in ONE call, 14 loose props → 8 typed
    (semantic filters nested in `filter`); scope='details' absorbs
    get_model_details; scope='graph' → kukai.query.graph_api (lazy)."""
    return {
        "type": "function",
        "function": {
            "name": "query_model",
            "description": (
                "Найти / посчитать / агрегировать элементы по ДЕКЛАРАТИВНОМУ "
                "фильтру БЕЗ написания C# — бэкенд исполняет проверенный "
                "version-safe шаблон (надёжнее execute_revit_code для поиска). "
                "НЕСКОЛЬКО категорий за ОДИН вызов (categories) — не вызывай "
                "инструмент по разу на категорию. scope: 'summary' (метрики из "
                "summary_of), 'elements' (ids), 'table' (строка на элемент — "
                "вместо C# со Select/OrderBy/Take), 'details' (детальный паспорт "
                "модели: типы/параметры/виды — раздел в section), 'graph' "
                "(запрос к графу модели). Точные имена типов бери из паспорта "
                "(СЛОВАРЬ) → filter.types."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "categories": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": (
                            "Категории (RU/EN алиасы: стены/двери/walls/…). "
                            "Несколько — за один вызов."
                        ),
                    },
                    "scope": {
                        "type": "string",
                        "enum": ["summary", "elements", "table", "details", "graph"],
                        "description": "Что вернуть. Default: summary.",
                    },
                    "table": {
                        "type": "object",
                        "description": (
                            "Для scope='table'. fields (≤12): id, name, "
                            "category, type, level, mark, area_m2, volume_m3, "
                            "length_m, height_mm, width_mm, 'param:<Имя>'. "
                            "order_by из fields; order asc|desc."
                        ),
                        "properties": {
                            "fields": {"type": "array", "items": {"type": "string"}},
                            "order_by": {"type": "string"},
                            "order": {"type": "string", "enum": ["asc", "desc"]},
                        },
                    },
                    "filter": {
                        "type": "object",
                        "description": "Декларативный фильтр элементов (все поля опциональны).",
                        "properties": {
                            "types": {
                                "type": "object",
                                "properties": {
                                    "contains": {"type": "string", "description": "Подстрока имени типа."},
                                    "names": {"type": "array", "items": {"type": "string"},
                                              "description": "Точные имена типов (самый надёжный фильтр)."},
                                },
                            },
                            "param": {
                                "type": "object",
                                "properties": {
                                    "name": {"type": "string"},
                                    "op": {"type": "string",
                                           "enum": ["empty", "not_empty", "eq", "contains", "gt", "lt"]},
                                    "value": {"type": "string"},
                                },
                                "description": "Фильтр по параметру экземпляра.",
                            },
                            "function": {
                                "type": "string",
                                "enum": ["exterior", "interior", "foundation",
                                         "retaining", "coreshaft", "soffit"],
                                "description": "Функция стены (WallType.Function): наружная=exterior, перегородка=interior.",
                            },
                            "width_mm": {
                                "type": "object",
                                "properties": {
                                    "op": {"type": "string", "enum": ["gt", "lt", "eq", "range"]},
                                    "value": {"type": "number"},
                                    "value2": {"type": "number"},
                                },
                                "description": "Реальная толщина стены (WallType.Width), мм.",
                            },
                            "material_contains": {
                                "type": "string",
                                "description": "Подстрока имени материала слоя (монолит/ЖБ → 'Железобетон').",
                            },
                            "level": {"type": "string", "description": "Подстрока имени уровня, напр. '03'."},
                            "selected": {"type": "boolean", "description": "Только текущее выделение."},
                        },
                    },
                    "summary_of": {
                        "type": "array",
                        "items": {"type": "string",
                                  "enum": ["count", "area_m2", "volume_m3", "by_type",
                                           "by_level", "coverage"]},
                        "description": (
                            "Метрики для scope='summary' (default ['count']). "
                            "by_type/by_level = разбивка по типам/этажам; добавь "
                            "area_m2/volume_m3 к by_level, чтобы получить площадь/"
                            "объём ПО КАЖДОМУ уровню одним вызовом (без C#). "
                            "coverage = заполненность filter.param.name за один вызов."
                        ),
                    },
                    "section": {
                        "type": "string",
                        "enum": ["full", "structure", "parameters", "spatial", "views", "standards"],
                        "description": "Раздел паспорта для scope='details'. Default: full.",
                    },
                    "graph": {
                        "type": "object",
                        "properties": {
                            "op": {"type": "string", "description": (
                                "Операция графа. Связи (хостинг/границы помещений): "
                                "hosted (двери/окна по стенам-хостам, args "
                                "{cat:door|window, host_type?, level?}); "
                                "room_boundaries (ограждающие элементы помещения, "
                                "args {room:имя|id}); rooms_without (помещения, ни "
                                "одна стена которых не несёт cat, args "
                                "{cat:window|door})."
                            )},
                            "args": {"type": "object", "description": "Аргументы операции."},
                        },
                        "description": "Запрос для scope='graph'.",
                    },
                    "action": {
                        "type": "string",
                        "enum": ["none", "select", "isolate", "highlight"],
                        "description": "Действие над найденными (default none).",
                    },
                    "limit": {
                        "type": "integer",
                        "description": "Предел ids для scope='elements' (default 1000).",
                    },
                },
                "required": [],
            },
        },
    }


def _get_family_tool_definitions() -> list[dict[str, Any]]:
    """10 purpose-built family-editor tools (V2 architecture).

    Each maps to a handler in kukai.llm.tool_handlers.family_tools that
    generates compliant C# from verified-API templates and dispatches
    through the standard bridge.execute path.
    """
    return [
        {
            "type": "function",
            "function": {
                "name": "inspect_family",
                "description": (
                    "Read-only snapshot of the active family doc — returns category, "
                    "template, all types with current parameter values, all parameters "
                    "with metadata, all solids (id+kind+bbox+subcategory+material), "
                    "all reference planes, all labeled dimensions, and all materials. "
                    "ALWAYS call this FIRST in any new family-editor conversation so you "
                    "know what's already in the family before adding/modifying anything."
                ),
                "parameters": {"type": "object", "properties": {}},
            },
        },
        {
            "type": "function",
            "function": {
                "name": "family_add_parameter",
                "description": (
                    "Add a FamilyParameter to the current family. Idempotent — skips if "
                    "a parameter with the same name already exists. Use BEFORE creating "
                    "types so the parameter is available for value-setting."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "name": {"type": "string", "description": "Parameter name, e.g. 'Width', 'Height', 'Material'."},
                        "spec_type": {
                            "type": "string",
                            "description": "Value type: Length|Area|Volume|Angle|Number|Integer|Boolean|Text|Material",
                            "enum": ["Length", "Area", "Volume", "Angle", "Number", "Integer", "Boolean", "Text", "Material"],
                        },
                        "group": {
                            "type": "string",
                            "description": "UI group: Dimensions|Constraints|Identity|Materials|Visibility|Structural|Mechanical|Electrical",
                            "enum": ["Dimensions", "Constraints", "Identity", "Materials", "Visibility", "Structural", "Mechanical", "Electrical"],
                        },
                        "is_instance": {"type": "boolean", "description": "true = per-instance, false = per-type (default)."},
                    },
                    "required": ["name"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "family_new_type",
                "description": (
                    "Create a new FamilyType in the current family and set its parameter values. "
                    "Idempotent — if a type with this name exists, it's updated instead of duplicated. "
                    "Length-typed parameter values are interpreted as millimetres."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "name": {"type": "string", "description": "Type name, e.g. 'Large', 'Small_900x600'."},
                        "param_values": {
                            "type": "object",
                            "description": (
                                "Dictionary {paramName: value}. Length values in millimetres. "
                                "Example: {\"Width\": 600, \"Height\": 800, \"Mark\": \"DR-001\"}"
                            ),
                        },
                    },
                    "required": ["name"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "family_extrude",
                "description": (
                    "Create a rectangular Extrusion solid at given coordinates (free, not bound "
                    "to a skeleton). Use for static panels: chair seat, table top, cabinet body. "
                    "If user wants the geometry to FLEX with parameters, first compose a skeleton "
                    "from primitives (family_add_parameter + family_create_reference_plane + "
                    "family_regenerate + family_create_dimension), then use this for solid bodies "
                    "and lock their faces via family_create_alignment."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "width_mm": {"type": "number", "description": "Rectangle X size in mm."},
                        "depth_mm": {"type": "number", "description": "Rectangle Y size in mm."},
                        "thickness_mm": {"type": "number", "description": "Extrusion thickness (Z) in mm."},
                        "z_offset_mm": {"type": "number", "description": "Bottom face Z position in mm (default 0)."},
                        "subcategory": {"type": "string", "description": "Optional subcategory name to assign to the solid."},
                    },
                    "required": ["width_mm", "depth_mm", "thickness_mm"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "family_cylinder",
                "description": (
                    "Create a cylinder extrusion — chair leg, post, column, lamp base. "
                    "Profile is a closed circle (two semicircular arcs). Coords in mm in family local frame."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "center_x_mm": {"type": "number", "description": "Cylinder centre X in mm."},
                        "center_y_mm": {"type": "number", "description": "Cylinder centre Y in mm."},
                        "radius_mm": {"type": "number", "description": "Radius in mm."},
                        "height_mm": {"type": "number", "description": "Cylinder height (Z) in mm."},
                        "z_offset_mm": {"type": "number", "description": "Base Z position in mm (default 0)."},
                        "subcategory": {"type": "string", "description": "Optional subcategory name."},
                    },
                    "required": ["radius_mm", "height_mm"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "family_void_cut",
                "description": (
                    "Drill a hole through existing solids via void extrusion. Works in "
                    "standard family templates (Generic Model, Furniture, Door, Window) — "
                    "voids auto-cut overlapping solids on regenerate, NO explicit "
                    "SolidSolidCutUtils call. shape='circle' uses radius_mm; "
                    "shape='rectangle' uses width_mm + depth_mm_xy."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "shape": {"type": "string", "enum": ["circle", "rectangle"], "description": "Void profile shape."},
                        "center_x_mm": {"type": "number", "description": "Void centre X in mm."},
                        "center_y_mm": {"type": "number", "description": "Void centre Y in mm."},
                        "z_offset_mm": {"type": "number", "description": "Void bottom Z in mm (start below solid for clean cut)."},
                        "depth_mm": {"type": "number", "description": "Void depth in mm (extends in +Z)."},
                        "radius_mm": {"type": "number", "description": "Required if shape='circle'."},
                        "width_mm": {"type": "number", "description": "Required if shape='rectangle' (X size)."},
                        "depth_mm_xy": {"type": "number", "description": "Required if shape='rectangle' (Y size)."},
                    },
                    "required": ["shape", "depth_mm"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "family_assign_material",
                "description": (
                    "Assign a Material (looked up by name, case-insensitive) to family solids "
                    "via BuiltInParameter.MATERIAL_ID_PARAM. target='all' assigns to every solid; "
                    "target=<subcategory name> assigns only to solids with that subcategory."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "material_name": {"type": "string", "description": "Material name (e.g. 'Wood - Oak', 'Metal - Steel')."},
                        "target": {"type": "string", "description": "'all' or a subcategory name to filter solids."},
                    },
                    "required": ["material_name"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "family_create_subcategory",
                "description": (
                    "Create a subcategory under the family's owner category — controls "
                    "per-part visibility and line weight. Idempotent."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "name": {"type": "string", "description": "Subcategory name, e.g. 'Frame', 'Glass', 'Seat'."},
                    },
                    "required": ["name"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "family_extrude_polygon",
                "description": (
                    "Extrude an arbitrary 2D polygon profile (any shape). Use for non-rectangular "
                    "panels, custom plates, L-shapes, hex profiles, anything closed-loop. "
                    "Profile auto-closes (last point links back to first)."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "points_mm": {
                            "type": "array",
                            "items": {"type": "array", "items": {"type": "number"}},
                            "description": "List of [x_mm, y_mm] vertices (>=3). Profile is auto-closed.",
                        },
                        "depth_mm": {"type": "number", "description": "Extrusion depth (Z) in mm."},
                        "z_offset_mm": {"type": "number", "description": "Bottom face Z position in mm (default 0)."},
                        "is_solid": {"type": "boolean", "description": "true=solid, false=void (auto-cuts overlapping solids)."},
                        "subcategory": {"type": "string", "description": "Optional subcategory name."},
                    },
                    "required": ["points_mm", "depth_mm"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "family_blend",
                "description": (
                    "Tapered solid blending between two arbitrary polygon profiles at different Z heights. "
                    "Use for lamp shades, tapered legs, prism transitions, frustums."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "bottom_points_mm": {
                            "type": "array", "items": {"type": "array", "items": {"type": "number"}},
                            "description": "Bottom profile [[x,y], ...]",
                        },
                        "top_points_mm": {
                            "type": "array", "items": {"type": "array", "items": {"type": "number"}},
                            "description": "Top profile [[x,y], ...]",
                        },
                        "bottom_z_mm": {"type": "number", "description": "Z of bottom profile (default 0)."},
                        "top_z_mm": {"type": "number", "description": "Z of top profile (default 300)."},
                        "is_solid": {"type": "boolean", "description": "true=solid, false=void."},
                    },
                    "required": ["bottom_points_mm", "top_points_mm"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "family_revolve",
                "description": (
                    "Revolution solid — rotate a 2D profile around an axis. Use for vases, bowls, "
                    "domes, spheres, bottles, columns with profiled silhouettes. Profile in XZ "
                    "plane (Y=0), must form a closed loop INCLUDING the axis-edge segment."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "profile_points_mm": {
                            "type": "array", "items": {"type": "array", "items": {"type": "number"}},
                            "description": "Closed loop of [x_mm, z_mm] points in XZ plane (Y=0 implicit).",
                        },
                        "axis_start_mm": {"type": "array", "items": {"type": "number"}, "description": "[x,y,z] start of rotation axis."},
                        "axis_end_mm": {"type": "array", "items": {"type": "number"}, "description": "[x,y,z] end of rotation axis."},
                        "start_angle_rad": {"type": "number", "description": "Start angle (radians). Default 0."},
                        "end_angle_rad": {"type": "number", "description": "End angle (radians). Default 2*PI."},
                        "is_solid": {"type": "boolean", "description": "true=solid, false=void."},
                    },
                    "required": ["profile_points_mm"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "family_delete_element",
                "description": "Delete a family element by ElementId. Use to clean up unwanted geometry.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "element_id": {"type": "integer", "description": "ElementId.Value from inspect_family."},
                    },
                    "required": ["element_id"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "family_move_element",
                "description": "Translate an existing element by [dx, dy, dz] in mm.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "element_id": {"type": "integer", "description": "ElementId.Value of element to move."},
                        "dx_mm": {"type": "number", "description": "X translation in mm."},
                        "dy_mm": {"type": "number", "description": "Y translation in mm."},
                        "dz_mm": {"type": "number", "description": "Z translation in mm."},
                    },
                    "required": ["element_id"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "family_create_reference_plane",
                "description": (
                    "Create a single named ReferencePlane — building block for parametric skeleton. "
                    "Use bubble→free direction for plane orientation, cut_dir as third basis (e.g. "
                    "BasisZ=(0,0,1) for vertical planes, BasisY=(0,1,0) for horizontal planes)."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "bubble_mm": {"type": "array", "items": {"type": "number"}, "description": "[x,y,z] bubble end."},
                        "free_mm": {"type": "array", "items": {"type": "number"}, "description": "[x,y,z] free end."},
                        "cut_dir": {"type": "array", "items": {"type": "number"}, "description": "[x,y,z] cut vector (e.g. [0,0,1] for vertical plane)."},
                        "name": {"type": "string", "description": "Optional name (e.g. 'Left', 'Center (Front/Back)')."},
                    },
                    "required": ["bubble_mm", "free_mm", "cut_dir"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "family_regenerate",
                "description": (
                    "Force doc.Regenerate(). REQUIRED between creating reference planes and "
                    "dimensioning them via family_create_dimension (Revit 2023+ regression: "
                    "'Invalid number of references' without this regenerate)."
                ),
                "parameters": {"type": "object", "properties": {}},
            },
        },
        {
            "type": "function",
            "function": {
                "name": "family_create_dimension",
                "description": (
                    "Create a labeled Dimension between two reference planes, bind to a FamilyParameter. "
                    "This is HOW parametric flex works: ref planes constrained by labeled dims that "
                    "drive a parameter value. Workflow: family_add_parameter → family_create_reference_plane "
                    "(twice) → family_regenerate → family_create_dimension."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "ref_plane_id_a": {"type": "integer", "description": "ElementId of first ref plane."},
                        "ref_plane_id_b": {"type": "integer", "description": "ElementId of second ref plane."},
                        "dim_line_p1_mm": {"type": "array", "items": {"type": "number"}, "description": "[x,y,z] dim line endpoint 1 (must lie in active view plane)."},
                        "dim_line_p2_mm": {"type": "array", "items": {"type": "number"}, "description": "[x,y,z] dim line endpoint 2."},
                        "family_param_name": {"type": "string", "description": "Name of existing FamilyParameter to bind (must be created first)."},
                    },
                    "required": ["ref_plane_id_a", "ref_plane_id_b"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "family_set_parameter_value",
                "description": (
                    "Set a single FamilyParameter value on the CURRENT type. For Length-typed "
                    "params, value is in mm. For text params pass string. For Yes/No pass boolean."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "param_name": {"type": "string", "description": "FamilyParameter name."},
                        "value": {
                            "description": "Value (number in mm for Length, string for Text, boolean for YesNo).",
                        },
                    },
                    "required": ["param_name", "value"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "family_create_model_lines",
                "description": "Create 3D model lines (visible in ALL view types). Use for locator/reference geometry inside family.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "lines_mm": {
                            "type": "array",
                            "items": {"type": "array", "items": {"type": "array", "items": {"type": "number"}}},
                            "description": "List of [[x1,y1,z1], [x2,y2,z2]] segments.",
                        },
                        "sketch_plane_normal": {"type": "array", "items": {"type": "number"}, "description": "[nx,ny,nz] plane normal (default [0,0,1])."},
                        "sketch_plane_origin_mm": {"type": "array", "items": {"type": "number"}, "description": "[x,y,z] plane origin in mm."},
                    },
                    "required": ["lines_mm"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "family_create_symbolic_lines",
                "description": (
                    "2D symbolic lines — visible ONLY in plan/elevation views (not 3D). "
                    "Used for УГО/СПДС schematic symbols on family annotations."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "lines_mm": {
                            "type": "array",
                            "items": {"type": "array", "items": {"type": "array", "items": {"type": "number"}}},
                            "description": "List of [[x1,y1,z1], [x2,y2,z2]] segments.",
                        },
                        "sketch_plane_normal": {"type": "array", "items": {"type": "number"}, "description": "[nx,ny,nz] plane normal (default [0,0,1])."},
                        "sketch_plane_origin_mm": {"type": "array", "items": {"type": "number"}, "description": "[x,y,z] plane origin in mm."},
                    },
                    "required": ["lines_mm"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "family_set_visibility",
                "description": (
                    "Set per-view-direction + per-detail-level visibility on a family solid. "
                    "Use for plan-only 2D symbols vs. 3D-only solid pairs (common LOD pattern)."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "element_id": {"type": "integer", "description": "Element id of the GenericForm (from inspect_family)."},
                        "in_plan": {"type": "boolean", "description": "Visible in plan/RCP views (default true)."},
                        "in_front_back": {"type": "boolean", "description": "Visible in front/back elevations."},
                        "in_left_right": {"type": "boolean", "description": "Visible in left/right elevations."},
                        "in_coarse": {"type": "boolean", "description": "Visible at coarse detail level."},
                        "in_medium": {"type": "boolean", "description": "Visible at medium detail level."},
                        "in_fine": {"type": "boolean", "description": "Visible at fine detail level."},
                    },
                    "required": ["element_id"],
                },
            },
        },
        # ─── V3: high-level composition primitives ────────────────────────
        {
            "type": "function",
            "function": {
                "name": "family_extrude_advanced",
                "description": (
                    "Powerful unified extrude — arbitrary profile (Lines + Arcs, "
                    "multi-loop with HOLES inside the outer boundary), arbitrary sketch "
                    "plane (axis-aligned OR INCLINED). Use this when the simple shortcuts "
                    "(family_extrude for rectangle, family_cylinder for circle, "
                    "family_extrude_polygon for line-only polygons) are not enough. "
                    "Examples: rings/washers (outer circle + inner-hole circle), gears "
                    "(involute teeth as composed arcs), L/U/hex with rounded corners, "
                    "panels on tilted surfaces. The Python validator REJECTS unclosed "
                    "loops and tells you which segment is bad — fix the gap and retry."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "profile": {
                            "type": "object",
                            "description": (
                                "{outer_loop: [<segment>, ...], inner_loops: [[...], ...]}. "
                                "outer_loop is required and must be a CLOSED chain (last segment "
                                "end == first segment start within 1 µm). inner_loops are HOLES "
                                "(optional), each closed, placed inside outer_loop.\n\n"
                                "Segment shapes:\n"
                                "  {\"type\": \"line\", \"p1\": [x_mm, y_mm], \"p2\": [x_mm, y_mm]}\n"
                                "  {\"type\": \"arc\",  \"center\": [cx_mm, cy_mm], \"radius\": r_mm,\n"
                                "                       \"start_deg\": <deg>, \"end_deg\": <deg>}\n"
                                "  {\"type\": \"arc\",  \"p1\": [x,y], \"p2\": [x,y], \"p3\": [x,y]}  # p2 on arc\n\n"
                                "All coords are LOCAL 2D mm in the sketch plane's UV frame.\n"
                                "Full-circle arcs (start_deg=0, end_deg=360) are auto-split into "
                                "two 180° halves (Revit rejects full-circle in a profile loop)."
                            ),
                            "properties": {
                                "outer_loop": {
                                    "type": "array",
                                    "items": {"type": "object"},
                                    "description": "Closed outer boundary as list of segments.",
                                },
                                "inner_loops": {
                                    "type": "array",
                                    "items": {"type": "array", "items": {"type": "object"}},
                                    "description": "Optional list of hole loops (each closed).",
                                },
                            },
                            "required": ["outer_loop"],
                        },
                        "sketch_plane": {
                            "type": "object",
                            "description": (
                                "Sketch plane spec: {origin_mm: [x,y,z], normal: [nx,ny,nz]}. "
                                "Default = XY plane at world origin (normal=[0,0,1], origin=[0,0,0]). "
                                "For inclined surfaces (sloped roof feature, half-timber diagonal), "
                                "supply a non-axis-aligned normal."
                            ),
                            "properties": {
                                "origin_mm": {"type": "array", "items": {"type": "number"}},
                                "normal": {"type": "array", "items": {"type": "number"}},
                            },
                        },
                        "depth_mm": {"type": "number", "description": "Extrusion length along plane normal in mm. Must be > 0."},
                        "is_solid": {"type": "boolean", "description": "true=solid (default), false=void (auto-cuts overlapping solids)."},
                        "subcategory": {"type": "string", "description": "Optional subcategory name."},
                    },
                    "required": ["profile", "depth_mm"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "family_sweep",
                "description": (
                    "Solid swept along a 2D path with a constant cross-section. Use for "
                    "handrails, pipes/conduit, gutters, decorative trim, torus (circular "
                    "path + circular profile), helices. Path is a list of Line/Arc segments; "
                    "profile is a closed loop of Line/Arc segments in the XY plane (Z=0 "
                    "implicit — Revit transforms internally)."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path_curves": {
                            "type": "array", "items": {"type": "object"},
                            "description": (
                                "Path segments (list of {type, ...}). Use 3D XYZ for line p1/p2 "
                                "or center/radius/angles for planar arcs. Sequential segments "
                                "share endpoints. Example torus path: one arc "
                                "{type:arc, center:[0,0,0], radius:200, start_deg:0, end_deg:360} "
                                "(auto-split into 2 halves)."
                            ),
                        },
                        "profile_loop": {
                            "type": "array", "items": {"type": "object"},
                            "description": (
                                "Closed 2D cross-section loop (Lines + Arcs). All coords are "
                                "[x_mm, y_mm] in the profile's own XY plane (Z=0 implicit). "
                                "Loop must be closed (first/last endpoints match within 1 µm)."
                            ),
                        },
                        "is_solid": {"type": "boolean", "description": "true=solid (default), false=void."},
                        "subcategory": {"type": "string", "description": "Optional subcategory name."},
                    },
                    "required": ["path_curves", "profile_loop"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "family_swept_blend",
                "description": (
                    "Solid swept along a SINGLE-curve path, morphing between two different "
                    "end profiles. Use for HVAC reducers (round-to-square), pipe transitions, "
                    "smooth decorative transitions. Path is EXACTLY ONE Line or Arc."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path_curve": {
                            "type": "object",
                            "description": (
                                "ONE segment dict (NOT a list). {type: 'line', p1: [...], p2: [...]} "
                                "or {type: 'arc', center: [...], radius: ..., start_deg: a, end_deg: b}."
                            ),
                        },
                        "start_profile": {
                            "type": "array", "items": {"type": "object"},
                            "description": "Closed loop of segments — bottom profile (all Z=0).",
                        },
                        "end_profile": {
                            "type": "array", "items": {"type": "object"},
                            "description": "Closed loop of segments — top profile (all Z=0).",
                        },
                        "is_solid": {"type": "boolean", "description": "true=solid (default), false=void."},
                    },
                    "required": ["path_curve", "start_profile", "end_profile"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "family_create_alignment",
                "description": (
                    "Lock alignment between two references — extrusion face ↔ reference plane, "
                    "or two reference planes. This is HOW parametric flex works: when the "
                    "reference plane moves (because its labeled dimension changed), the locked "
                    "face follows. NewAlignment LOCKS an existing geometric alignment — the "
                    "geometry must ALREADY coincide before calling. View must be 2D (plan/elevation)."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "anchor": {
                            "type": "object",
                            "description": (
                                "Anchor reference. Either:\n"
                                "  {\"type\": \"reference_plane\", \"id\": <int>}\n"
                                "  {\"type\": \"extrusion_face\", \"element_id\": <int>, \"face_normal\": [nx,ny,nz]}"
                            ),
                        },
                        "target": {
                            "type": "object",
                            "description": "Target reference — same shape as anchor.",
                        },
                    },
                    "required": ["anchor", "target"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "family_create_array",
                "description": (
                    "Replicate one or more existing solids N times — radial (gear teeth, "
                    "fence pickets around a circle, ferris-wheel cabins) or linear (stair "
                    "treads, balusters, repeating panel features). NOT a Revit parametric "
                    "array — deterministic CopyElement+RotateElement in a single transaction. "
                    "Use after first creating ONE source instance via family_extrude / "
                    "family_cylinder etc., then call this to multiply. Source element IDs "
                    "come from the return value of the creation tool or from inspect_family."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "source_element_ids": {
                            "type": "array",
                            "items": {"type": "integer"},
                            "description": "ElementId.Value of solids to replicate. Pass a list — for a single solid pass [id].",
                        },
                        "array_type": {
                            "type": "string",
                            "enum": ["radial", "linear"],
                            "description": "radial = rotate copies around an axis; linear = translate copies along a vector.",
                        },
                        "count": {
                            "type": "integer",
                            "description": "Total copies INCLUDING the source (e.g. 20 gear teeth = count=20). Min 2, max 200.",
                        },
                        "rotation_center_mm": {
                            "type": "array", "items": {"type": "number"},
                            "description": "[x,y,z] in mm — pivot point for radial array (required for radial).",
                        },
                        "rotation_axis": {
                            "type": "array", "items": {"type": "number"},
                            "description": "[nx,ny,nz] direction of rotation axis. Default [0,0,1] (Z axis). Used for radial.",
                        },
                        "total_angle_deg": {
                            "type": "number",
                            "description": (
                                "Total sweep angle in degrees. 360 = full ring with even spacing "
                                "(N copies, last wraps to start). For partial arcs (e.g. 180) the "
                                "copies are spread end-to-end. Default 360."
                            ),
                        },
                        "translation_step_mm": {
                            "type": "array", "items": {"type": "number"},
                            "description": "[dx,dy,dz] in mm — offset per step for linear array (required for linear).",
                        },
                    },
                    "required": ["source_element_ids", "array_type", "count"],
                },
            },
        },
        # ─── V4: Code-CAD "any complexity" path ───────────────────────────
        {
            "type": "function",
            "function": {
                "name": "family_generate_complex",
                "description": (
                    "Generate ANY-complexity geometry by writing CadQuery Python code on the "
                    "server. Use THIS tool — not V3 primitives — for organic / freeform / "
                    "detailed solids that exceed what box / cylinder / sweep / blend / revolve "
                    "can express: smooth car bodies, statues, decorative carvings, organic "
                    "wheel arches, custom mechanical parts with fillets / chamfers / NURBS "
                    "surfaces, lofted shapes through arbitrary cross-sections.\n\n"
                    "The server runs your CadQuery code in a sandboxed Python subprocess, "
                    "exports an STL, and imports it into the family as a single DirectShape. "
                    "Result: real curved geometry, NOT a stack of boxes.\n\n"
                    "Code conventions:\n"
                    "  • `cadquery as cq` is pre-imported. Don't re-import.\n"
                    "  • Assign final result to a variable named `result` (or model/final/"
                    "output/shape). Example:\n"
                    "      result = (cq.Workplane('XY')\n"
                    "          .box(1600, 800, 350)\n"
                    "          .edges('|Z').fillet(80)\n"
                    "          .edges('|Z and >X').chamfer(50, 30))\n"
                    "  • All numeric units are in MILLIMETRES (mm). Server converts to feet.\n"
                    "  • Use unions / cuts / intersections for composite shapes:\n"
                    "      body = chassis.union(cabin).union(hood)\n"
                    "      result = body.cut(wheel_well_left).cut(wheel_well_right)\n\n"
                    "Multi-part result (RECOMMENDED for complex objects):\n"
                    "  Use `cq.Assembly()` for objects with multiple coloured parts —\n"
                    "  the tool creates ONE DirectShape per part, each with the\n"
                    "  Assembly child's colour applied.\n"
                    "    asm = cq.Assembly()\n"
                    "    asm.add(body,   name='body',   color=cq.Color(0.1, 0.1, 0.5))   # dark blue\n"
                    "    asm.add(wheels, name='wheels', color=cq.Color(0.05, 0.05, 0.05))  # near-black\n"
                    "    result = asm\n\n"
                    "Response includes: element_ids list, bbox_mm (auto-parametrize on this),\n"
                    "parts_summary (per-part triangle count + colour), and an SVG preview\n"
                    "of the geometry that you can use to verify the shape on the next round.\n\n"
                    "Trade-off vs V3 primitives: result is static DirectShape (NO labeled-"
                    "dimension flex). Pick this tool when curved / complex / multi-coloured "
                    "geometry is the requirement; pick V3 primitives when parametric "
                    "Width / Depth / Height flex is needed."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "code": {
                            "type": "string",
                            "description": (
                                "CadQuery Python source. MUST assign the final shape to a "
                                "variable named `result` (or model/final/output/shape). All "
                                "numeric values in millimetres."
                            ),
                        },
                        "description": {
                            "type": "string",
                            "description": "Short label (max 120 chars) for the resulting DirectShape, e.g. 'rolls_royce_body'.",
                        },
                        "tolerance": {
                            "type": "number",
                            "description": (
                                "STL linear tolerance in mm. Controls mesh smoothness vs triangle count.\n"
                                "  • 0.5-1.0 — coarse, fast, ~hundreds of triangles. For block-like primitives.\n"
                                "  • 0.1 (default) — balanced, ~thousands of triangles.\n"
                                "  • 0.02-0.05 — HIGH QUALITY, 5-10x more triangles, NEAR-NURBS smoothness "
                                "for car bodies, sculptures, ornamental work. Use this for organic curves "
                                "where triangulation artifacts would be visible. Stay under 50K total "
                                "triangles (the response carries triangle_count — pull tolerance back up "
                                "if you hit the oversized error)."
                            ),
                        },
                        "quality": {
                            "type": "string",
                            "enum": ["draft", "balanced", "high"],
                            "description": (
                                "Convenience shortcut for tolerance. Overrides explicit tolerance if set.\n"
                                "  • 'draft' = tolerance 0.5 (fast, blocky — for prototyping shapes)\n"
                                "  • 'balanced' (default) = tolerance 0.1\n"
                                "  • 'high' = tolerance 0.03 (smooth curves — use for organic / showcase work)"
                            ),
                        },
                        "angular_tolerance": {
                            "type": "number",
                            "description": "STL angular tolerance in radians. Default 0.1.",
                        },
                        "timeout_s": {
                            "type": "number",
                            "description": "Subprocess timeout in seconds. Default 60. Raise to 300 for very complex models.",
                        },
                        "output_format": {
                            "type": "string",
                            "enum": ["stl", "step", "auto"],
                            "description": (
                                "Geometry transport format.\n"
                                "  • 'stl' (default) — fast, every Revit version, but FACETED "
                                "(visible triangles on curves). Use for parametric primitives, "
                                "mechanical parts where facets are acceptable.\n"
                                "  • 'step' — preserves NURBS curves losslessly (smooth car "
                                "bodies, ornamental surfaces). Revit 2024+. Slower import. Use "
                                "for organic / curved bodies where smoothness matters.\n"
                                "  • 'auto' — request STEP, fall back to STL if bridge doesn't "
                                "support it. Safest choice for high-quality work."
                            ),
                        },
                    },
                    "required": ["code"],
                },
            },
        },
    ]
