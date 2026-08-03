"""Excel operations — Gemini controls Excel via commands, not raw data.

Supports: sort, filter, totals, formatting, pivot, formulas.
All operations work on existing Excel files or stored data.
"""

from __future__ import annotations

import logging
from io import BytesIO
from pathlib import Path
from typing import Any, Optional

logger = logging.getLogger(__name__)


def apply_excel_operations(
    wb_bytes: bytes,
    operations: list[dict[str, Any]],
) -> bytes:
    """Apply a list of operations to an Excel workbook.

    Args:
        wb_bytes: Raw bytes of an .xlsx file
        operations: List of operation dicts, each with "type" + params

    Returns:
        Modified .xlsx bytes
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(BytesIO(wb_bytes))
    ws = wb.active

    for op in operations:
        op_type = op.get("type", "")

        try:
            if op_type == "sort":
                _sort_sheet(ws, op)
            elif op_type == "auto_filter":
                _auto_filter(ws)
            elif op_type == "freeze_header":
                ws.freeze_panes = "A2"
            elif op_type == "column_widths":
                _auto_column_widths(ws, op)
            elif op_type == "header_format":
                _format_header(ws, op)
            elif op_type == "add_totals":
                _add_totals(ws, op)
            elif op_type == "add_formula_column":
                _add_formula_column(ws, op)
            elif op_type == "conditional_format":
                _conditional_format(ws, op)
            elif op_type == "pivot":
                _create_pivot_sheet(wb, ws, op)
            else:
                logger.warning("Unknown excel operation: %s", op_type)
        except Exception as e:
            logger.warning("Excel operation '%s' failed: %s", op_type, e)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def create_formatted_excel(
    data: list[dict[str, Any]],
    filename: str = "report.xlsx",
    sheet_name: str = "Report",
    sort_by: str = "",
    sort_order: str = "asc",
    add_totals: bool = False,
    total_columns: list[str] | None = None,
    auto_filter: bool = True,
    freeze_header: bool = True,
    header_color: str = "4472C4",
) -> bytes:
    """Create a well-formatted Excel file from data.

    This is the 'batteries-included' version — creates a professional-looking
    spreadsheet without Gemini needing to specify every detail.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    if not data:
        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    # Write headers
    headers = list(data[0].keys())
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Write data
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header, "")
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Sort
    if sort_by and sort_by in headers:
        sort_col = headers.index(sort_by) + 1
        data_rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False))
        reverse = sort_order.lower() in ("desc", "descending")

        def _sk(row):
            v = row[sort_col - 1].value
            if v is None:
                return (2, 0, "")
            if isinstance(v, (int, float)):
                return (0, v, "")
            return (1, 0, str(v))

        try:
            data_rows.sort(key=_sk, reverse=reverse)
            row_values = [[cell.value for cell in row] for row in data_rows]
            for new_row_idx, values in enumerate(row_values, 2):
                for col_idx2, v in enumerate(values, 1):
                    ws.cell(row=new_row_idx, column=col_idx2, value=v)
        except Exception:
            pass

    # Totals row
    if add_totals:
        total_row = ws.max_row + 1
        ws.cell(row=total_row, column=1, value="ИТОГО").font = Font(bold=True)
        cols_to_sum = total_columns or headers
        for col_idx, header in enumerate(headers, 1):
            if header in cols_to_sum:
                # Check if column is numeric
                sample = ws.cell(row=2, column=col_idx).value
                if isinstance(sample, (int, float)):
                    col_letter = get_column_letter(col_idx)
                    ws.cell(
                        row=total_row, column=col_idx,
                        value=f"=SUM({col_letter}2:{col_letter}{total_row - 1})",
                    ).font = Font(bold=True)

    # Auto filter
    if auto_filter and ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    # Freeze header
    if freeze_header:
        ws.freeze_panes = "A2"

    # Auto column widths
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 100), min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 50)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


# --- Internal operation handlers ---

def _sort_sheet(ws, op):
    """Sort worksheet by column."""
    column = op.get("column", "")
    order = op.get("order", "asc")
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    if column not in headers:
        return
    col_idx = headers.index(column) + 1
    data_rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False))
    reverse = order.lower() in ("desc", "descending")

    def _sort_key(row):
        v = row[col_idx - 1].value
        if v is None:
            return (2, 0, "")
        if isinstance(v, (int, float)):
            return (0, v, "")
        return (1, 0, str(v))

    data_rows.sort(key=_sort_key, reverse=reverse)
    # Extract values BEFORE writing — cells are references, not copies
    row_values = [[cell.value for cell in row] for row in data_rows]
    for new_row_idx, values in enumerate(row_values, 2):
        for c, v in enumerate(values, 1):
            ws.cell(row=new_row_idx, column=c, value=v)


def _auto_filter(ws):
    from openpyxl.utils import get_column_letter
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"


def _auto_column_widths(ws, op):
    from openpyxl.utils import get_column_letter
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 100), min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 50)


def _format_header(ws, op):
    from openpyxl.styles import Font, PatternFill, Alignment
    color = op.get("color", "4472C4")
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center")


def _add_totals(ws, op):
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    columns = op.get("columns", [])
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=1, value="ИТОГО").font = Font(bold=True)
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    for col_idx, header in enumerate(headers, 1):
        if not columns or header in columns:
            sample = ws.cell(row=2, column=col_idx).value
            if isinstance(sample, (int, float)):
                col_letter = get_column_letter(col_idx)
                ws.cell(
                    row=total_row, column=col_idx,
                    value=f"=SUM({col_letter}2:{col_letter}{total_row - 1})",
                ).font = Font(bold=True)


def _add_formula_column(ws, op):
    """Add a calculated column using Excel formula."""
    from openpyxl.utils import get_column_letter
    name = op.get("name", "Calculated")
    formula_template = op.get("formula", "")
    new_col = ws.max_column + 1
    ws.cell(row=1, column=new_col, value=name)
    for row in range(2, ws.max_row + 1):
        formula = formula_template.replace("{row}", str(row))
        ws.cell(row=row, column=new_col, value=formula)


def _conditional_format(ws, op):
    """Apply conditional formatting (highlight cells by condition)."""
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter
    column = op.get("column", "")
    operator = op.get("operator", "greaterThan")
    value = op.get("value", 0)
    color = op.get("color", "FF0000")
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    if column not in headers:
        return
    col_idx = headers.index(column) + 1
    col_letter = get_column_letter(col_idx)
    cell_range = f"{col_letter}2:{col_letter}{ws.max_row}"
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    ws.conditional_formatting.add(cell_range, CellIsRule(operator=operator, formula=[str(value)], fill=fill))


def _create_pivot_sheet(wb, source_ws, op):
    """Create a pivot-like summary sheet using Excel formulas (SUMIF/COUNTIF/AVERAGEIF).

    Values are LIVE Excel formulas by design: opening the report in Excel
    recalculates them, and editing the source data updates the pivot. openpyxl does
    not evaluate formulas, so a programmatic read-back sees the formula TEXT — tests
    assert the formula string, not a computed number.
    """
    from openpyxl.utils import get_column_letter

    rows_field = op.get("rows", "")
    values_field = op.get("values", "")
    aggfunc = op.get("aggfunc", "sum")
    sheet_name = op.get("sheet_name", "Pivot")

    headers = [source_ws.cell(row=1, column=c).value for c in range(1, source_ws.max_column + 1)]
    if rows_field not in headers or values_field not in headers:
        return

    rows_col = headers.index(rows_field) + 1
    values_col = headers.index(values_field) + 1

    # Collect unique keys from source sheet
    unique_keys = []
    seen: set[str] = set()
    for row in range(2, source_ws.max_row + 1):
        key = str(source_ws.cell(row=row, column=rows_col).value or "")
        if key and key not in seen:
            seen.add(key)
            unique_keys.append(key)
    unique_keys.sort()

    if not unique_keys:
        return

    # Build cross-sheet formula references
    source_name = source_ws.title
    rows_letter = get_column_letter(rows_col)
    values_letter = get_column_letter(values_col)
    max_row = source_ws.max_row
    range_rows = f"'{source_name}'!{rows_letter}$2:{rows_letter}${max_row}"
    range_vals = f"'{source_name}'!{values_letter}$2:{values_letter}${max_row}"

    # Create pivot sheet
    pivot_ws = wb.create_sheet(title=sheet_name)
    pivot_ws.cell(row=1, column=1, value=rows_field)
    agg_label = {"sum": "Сумма", "count": "Количество", "avg": "Среднее"}.get(aggfunc, aggfunc)
    pivot_ws.cell(row=1, column=2, value=f"{agg_label} ({values_field})")

    from openpyxl.styles import Font, PatternFill
    for col in range(1, 3):
        cell = pivot_ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Write unique keys + formula for each
    for row_idx, key in enumerate(unique_keys, 2):
        pivot_ws.cell(row=row_idx, column=1, value=key)
        ref = f"A{row_idx}"
        if aggfunc == "sum":
            formula = f"=SUMIF({range_rows},{ref},{range_vals})"
        elif aggfunc == "count":
            formula = f"=COUNTIF({range_rows},{ref})"
        elif aggfunc == "avg":
            formula = f"=AVERAGEIF({range_rows},{ref},{range_vals})"
        else:
            formula = f"=SUMIF({range_rows},{ref},{range_vals})"
        pivot_ws.cell(row=row_idx, column=2, value=formula)

    # ИТОГО row with SUM formula
    total_row = len(unique_keys) + 2
    total_cell = pivot_ws.cell(row=total_row, column=1, value="ИТОГО")
    total_cell.font = Font(bold=True)
    sum_cell = pivot_ws.cell(row=total_row, column=2, value=f"=SUM(B2:B{total_row - 1})")
    sum_cell.font = Font(bold=True)

    # Auto column widths
    for col in range(1, 3):
        max_len = max(
            len(str(pivot_ws.cell(row=r, column=col).value or ""))
            for r in range(1, total_row + 1)
        )
        pivot_ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 50)
