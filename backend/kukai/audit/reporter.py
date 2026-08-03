"""AuditReporter — formats audit results for chat and Excel export.

Two output modes:
- format_chat(): Markdown string for streaming into the chat UI
- generate_excel(): .xlsx bytes for file delivery to the client
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from kukai.audit.models import (
    AuditIssue,
    AuditReport,
    AuditStatus,
    Category,
    Severity,
)

# ---------------------------------------------------------------------------
# Severity display helpers
# ---------------------------------------------------------------------------

_SEVERITY_LABEL = {
    Severity.ERROR: "ОШИБКА",
    Severity.WARNING: "ПРЕДУПРЕЖДЕНИЕ",
    Severity.INFO: "ИНФО",
}

_STATUS_LABEL = {
    AuditStatus.PASSED: "ПРОЙДЕН",
    AuditStatus.PASSED_WITH_WARNINGS: "ПРОЙДЕН С ЗАМЕЧАНИЯМИ",
    AuditStatus.FAILED: "НЕ ПРОЙДЕН",
}

_CATEGORY_LABEL = {
    Category.NAMING: "Именование",
    Category.PARAMETERS: "Параметры",
    Category.GEOMETRY: "Геометрия",
    Category.LEVELS: "Уровни и привязки",
    Category.WARNINGS: "Предупреждения Revit",
    Category.IFC_READY: "Готовность к IFC",
}

# Excel styling constants
_HEADER_BG = "2F5496"
_HEADER_FG = "FFFFFF"
_ERROR_BG = "FFC7CE"
_WARNING_BG = "FFEB9C"
_INFO_BG = "D6E4F0"
_PASSED_BG = "C6EFCE"


class AuditReporter:
    """Formats AuditReport into chat messages and Excel files."""

    # ------------------------------------------------------------------
    # Chat (Markdown) output
    # ------------------------------------------------------------------

    @staticmethod
    def format_chat(report: AuditReport) -> str:
        """Format an AuditReport as a Markdown string for the chat UI."""
        s = report.summary
        status = _STATUS_LABEL.get(report.status, report.status.value)

        lines: list[str] = []
        lines.append(f"## Результат аудита: {status}")
        lines.append("")
        lines.append(
            f"Проверено правил: **{s.total_rules}** | "
            f"Ошибок: **{s.errors}** | "
            f"Предупреждений: **{s.warnings}** | "
            f"Инфо: **{s.infos}** | "
            f"Пройдено: **{s.passed}**"
        )
        lines.append(f"Элементов в модели: **{report.total_elements}** | "
                      f"Время: **{report.duration_seconds:.1f}с**")
        lines.append("")

        if not report.issues:
            lines.append("Все проверки пройдены успешно.")
            return "\n".join(lines)

        # Group issues by category
        by_cat: dict[Category, list[AuditIssue]] = {}
        for issue in report.issues:
            by_cat.setdefault(issue.category, []).append(issue)

        for cat in Category:
            issues = by_cat.get(cat)
            if not issues:
                continue
            cat_label = _CATEGORY_LABEL.get(cat, cat.value)
            lines.append(f"### {cat_label}")
            lines.append("")
            for issue in issues:
                sev = _SEVERITY_LABEL.get(issue.severity, issue.severity.value)
                lines.append(f"- **[{sev}]** {issue.rule_name} ({issue.rule_id})")
                if issue.message:
                    lines.append(f"  {issue.message}")
                if issue.element_count > 0:
                    lines.append(f"  Затронуто элементов: {issue.element_count}")
                if issue.explanation:
                    lines.append(f"  > {issue.explanation}")
                lines.append("")

        return "\n".join(lines)

    # ------------------------------------------------------------------
    # Excel output
    # ------------------------------------------------------------------

    @staticmethod
    def generate_excel(report: AuditReport) -> bytes:
        """Generate an Excel workbook from the audit report.

        Returns raw bytes suitable for WebSocket file delivery.
        Two sheets:
          1. "Сводка" — summary and status
          2. "Замечания" — detailed issue list
        """
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        thin = Side(style="thin")
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_font = Font(bold=True, color=_HEADER_FG, size=11)
        header_fill = PatternFill(start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        wrap_align = Alignment(vertical="top", wrap_text=True)

        sev_fills = {
            Severity.ERROR: PatternFill(start_color=_ERROR_BG, end_color=_ERROR_BG, fill_type="solid"),
            Severity.WARNING: PatternFill(start_color=_WARNING_BG, end_color=_WARNING_BG, fill_type="solid"),
            Severity.INFO: PatternFill(start_color=_INFO_BG, end_color=_INFO_BG, fill_type="solid"),
        }

        wb = Workbook()

        # --- Sheet 1: Summary ---
        ws_sum = wb.active
        ws_sum.title = "Сводка"
        ws_sum.sheet_properties.tabColor = "2F5496"

        status = _STATUS_LABEL.get(report.status, report.status.value)
        summary_rows = [
            ("Статус аудита", status),
            ("Всего правил", report.summary.total_rules),
            ("Пройдено", report.summary.passed),
            ("Ошибок", report.summary.errors),
            ("Предупреждений", report.summary.warnings),
            ("Информационных", report.summary.infos),
            ("Элементов в модели", report.total_elements),
            ("Время проверки (сек)", report.duration_seconds),
            ("Дата проверки", report.checked_at.strftime("%Y-%m-%d %H:%M:%S")),
            ("Категории", ", ".join(report.categories_checked)),
        ]

        for r, (label, value) in enumerate(summary_rows, start=1):
            cell_label = ws_sum.cell(row=r, column=1, value=label)
            cell_label.font = Font(bold=True)
            cell_label.border = thin_border
            cell_value = ws_sum.cell(row=r, column=2, value=value)
            cell_value.border = thin_border

        ws_sum.column_dimensions["A"].width = 25
        ws_sum.column_dimensions["B"].width = 50

        # Color the status cell
        status_fill_map = {
            AuditStatus.PASSED: PatternFill(start_color=_PASSED_BG, end_color=_PASSED_BG, fill_type="solid"),
            AuditStatus.PASSED_WITH_WARNINGS: PatternFill(start_color=_WARNING_BG, end_color=_WARNING_BG, fill_type="solid"),
            AuditStatus.FAILED: PatternFill(start_color=_ERROR_BG, end_color=_ERROR_BG, fill_type="solid"),
        }
        ws_sum.cell(row=1, column=2).fill = status_fill_map.get(
            report.status,
            PatternFill(fill_type=None),
        )

        # --- Sheet 2: Issues ---
        ws_iss = wb.create_sheet("Замечания")
        ws_iss.sheet_properties.tabColor = "ED7D31"

        headers = [
            "ID правила", "Правило", "Категория", "Серьёзность",
            "Сообщение", "Элементов", "ID элементов", "Объяснение",
        ]
        col_widths = [12, 30, 18, 16, 45, 12, 25, 50]

        for c, (header, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws_iss.cell(row=1, column=c, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws_iss.column_dimensions[get_column_letter(c)].width = width

        for r, issue in enumerate(report.issues, start=2):
            sev_label = _SEVERITY_LABEL.get(issue.severity, issue.severity.value)
            cat_label = _CATEGORY_LABEL.get(issue.category, issue.category.value)
            element_ids_str = ", ".join(str(eid) for eid in issue.element_ids[:50])
            if len(issue.element_ids) > 50:
                element_ids_str += f"... (+{len(issue.element_ids) - 50})"

            row_data = [
                issue.rule_id,
                issue.rule_name,
                cat_label,
                sev_label,
                issue.message,
                issue.element_count,
                element_ids_str,
                issue.explanation or "",
            ]

            sev_fill = sev_fills.get(issue.severity)
            for c, value in enumerate(row_data, start=1):
                cell = ws_iss.cell(row=r, column=c, value=value)
                cell.border = thin_border
                cell.alignment = wrap_align
                if sev_fill and c == 4:
                    cell.fill = sev_fill

        # Auto-filter on issues sheet
        if report.issues:
            last_col = get_column_letter(len(headers))
            ws_iss.auto_filter.ref = f"A1:{last_col}{len(report.issues) + 1}"

        # Freeze header row
        ws_iss.freeze_panes = "A2"

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()
